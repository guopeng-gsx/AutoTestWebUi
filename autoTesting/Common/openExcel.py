from openpyxl import load_workbook

wb = load_workbook(filename ='../Data/empty_book1.xlsx')
sheet_ranges = wb['range names']
dest_filename: str = '../Data/empty_book1.xlsx'
ws1 = wb.active
ws1.title = "range names"
sheet_ranges = wb['range names']
def getname():
    name = sheet_ranges.cell(2, 1).value
    return name
def getpwd():
    pwd = sheet_ranges.cell(2, 2).value
    return pwd